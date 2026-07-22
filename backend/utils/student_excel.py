"""Workbook persistence for student self-registration."""

from pathlib import Path
from threading import Lock

from openpyxl import load_workbook


_WORKBOOK_LOCK = Lock()
_WORKBOOK_PATH = Path(__file__).resolve().parents[1] / "scripts" / "Automated_Answer_Evaluation_Login_Data.xlsx"


def append_registered_student(*, full_name, email, password_hash):
    """Append a student record to the project workbook.

    Passwords are stored as hashes, never as plain text.  The application
    remains the source of truth for authentication in MongoDB.
    """
    with _WORKBOOK_LOCK:
        if not _WORKBOOK_PATH.exists():
            raise FileNotFoundError(f"Student workbook was not found: {_WORKBOOK_PATH}")

        workbook = load_workbook(_WORKBOOK_PATH)
        sheet = workbook["Students"]
        next_number = sheet.max_row
        student_id = f"STU{next_number:03d}"
        roll_no = f"REG{next_number:04d}"

        sheet.append([
            student_id,
            full_name,
            email,
            password_hash,
            "Not assigned",
            "Not assigned",
            "Not assigned",
            roll_no,
            "Active",
        ])
        workbook.save(_WORKBOOK_PATH)

    return {"studentId": student_id, "rollNo": roll_no}


def update_student_password(*, email, password_hash):
    """Update the matching student password hash in the workbook."""
    with _WORKBOOK_LOCK:
        if not _WORKBOOK_PATH.exists():
            raise FileNotFoundError(f"Student workbook was not found: {_WORKBOOK_PATH}")

        workbook = load_workbook(_WORKBOOK_PATH)
        sheet = workbook["Students"]
        headers = {cell.value: cell.column for cell in sheet[1]}
        email_column = headers.get("Email")
        password_column = headers.get("Password")
        if not email_column or not password_column:
            raise ValueError("Students workbook does not have Email and Password columns.")

        for row in range(2, sheet.max_row + 1):
            if str(sheet.cell(row, email_column).value or "").strip().lower() == email.lower():
                sheet.cell(row, password_column).value = password_hash
                workbook.save(_WORKBOOK_PATH)
                return True

    return False
