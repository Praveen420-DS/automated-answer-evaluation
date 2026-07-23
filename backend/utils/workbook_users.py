"""Read the supplied CSV/Excel login data when demo accounts are not in MongoDB."""

import csv
from pathlib import Path

from openpyxl import load_workbook
from werkzeug.security import check_password_hash, generate_password_hash


WORKBOOK_NAME = "Automated_Answer_Evaluation_Login_Data.xlsx"
WORKBOOK_PATHS = (
    Path(__file__).resolve().parents[1] / "data" / WORKBOOK_NAME,
    Path(__file__).resolve().parents[1] / "scripts" / WORKBOOK_NAME,
)
DATA_DIR = Path(__file__).resolve().parents[1] / "data"
CSV_SOURCES = (
    ("Automated_Answer_Evaluation_Login_Data-Students.csv", "student", "Student ID"),
    ("Automated_Answer_Evaluation_Login_Data-Teachers.csv", "faculty", "Teacher ID"),
    ("Automated_Answer_Evaluation_Login_Data-Admin.csv", "admin", "Admin ID"),
)


def _text(value):
    return str(value or "").strip()


def _matching_user(values, identifier, password, role, id_header):
    """Validate one normalized CSV/worksheet row and build an app user."""
    wanted = identifier.strip().lower()
    email = _text(values.get("Email")).lower()
    user_id = _text(values.get(id_header)).lower()
    if wanted not in {email, user_id, email.split("@", 1)[0]}:
        return None
    if _text(values.get("Status", "Active")).lower() != "active":
        return None
    saved_password = _text(values.get("Password"))
    valid_password = (
        check_password_hash(saved_password, password)
        if saved_password.startswith(("pbkdf2:", "scrypt:"))
        else saved_password == password
    )
    if not valid_password:
        return None
    user = {
        "fullName": _text(values.get("Name")), "email": email,
        "username": user_id or email.split("@", 1)[0],
        "password": generate_password_hash(password), "role": role,
        "status": "Active",
    }
    if role == "faculty":
        user.update({"facultyId": user_id, "department": _text(values.get("Department")), "designation": _text(values.get("Designation")), "subject": _text(values.get("Subject"))})
    elif role == "admin":
        user.update({"adminId": user_id, "adminRole": _text(values.get("Role"))})
    else:
        user.update({"studentId": user_id, "department": _text(values.get("Department")), "year": values.get("Year"), "section": _text(values.get("Section")), "rollNo": _text(values.get("Roll No"))})
    return user


def find_csv_user(identifier, password):
    """Find and validate Student, Staff, or Admin credentials from project CSVs."""
    for filename, role, id_header in CSV_SOURCES:
        path = DATA_DIR / filename
        if not path.exists():
            continue
        with path.open(encoding="utf-8-sig", newline="") as file:
            for row in csv.DictReader(file):
                user = _matching_user(row, identifier, password, role, id_header)
                if user:
                    return user
    return None


def find_workbook_user(identifier, password):
    """Return a normalized user record if an active workbook credential matches.

    The original demo workbook contains plain-text passwords.  They are used
    only for this one-time verification; the MongoDB copy is always saved with
    a password hash by the login route.
    """
    workbook_path = next((path for path in WORKBOOK_PATHS if path.exists()), None)
    if workbook_path is None:
        return None

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        for sheet_name, role, id_header in (
            ("Students", "student", "Student ID"),
            ("Teachers", "faculty", "Teacher ID"),
            ("Admin", "admin", "Admin ID"),
        ):
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            headers = {str(value).strip(): index for index, value in enumerate(next(sheet.iter_rows(values_only=True)), start=0) if value}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                values = {header: row[index] if index < len(row) else None for header, index in headers.items()}
                user = _matching_user(values, identifier, password, role, id_header)
                if user:
                    return user
    finally:
        workbook.close()
    return None
